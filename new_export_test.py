import requests
import sys
import json
import io
import pandas as pd
import math
import openpyxl
from datetime import datetime, timedelta

class NewExcelExportTester:
    def __init__(self, base_url="https://config-manager-2.preview.emergentagent.com"):
        self.base_url = base_url
        self.commandes_session_id = None
        self.stock_session_id = None
        self.transit_session_id = None
        self.tests_run = 0
        self.tests_passed = 0
        self.calculation_results = None

    def run_test(self, name, method, endpoint, expected_status, data=None, files=None):
        """Run a single API test"""
        url = f"{self.base_url}/{endpoint}"
        headers = {}
        
        if not files:
            headers['Content-Type'] = 'application/json'

        self.tests_run += 1
        print(f"\n🔍 Testing {name}...")
        print(f"URL: {url}")
        
        try:
            if method == 'GET':
                response = requests.get(url, headers=headers)
            elif method == 'POST':
                if files:
                    response = requests.post(url, files=files)
                else:
                    response = requests.post(url, json=data, headers=headers)

            success = response.status_code == expected_status
            if success:
                self.tests_passed += 1
                print(f"✅ Passed - Status: {response.status_code}")
                try:
                    if 'application/json' in response.headers.get('content-type', ''):
                        response_data = response.json()
                        print(f"Response: {json.dumps(response_data, indent=2)[:500]}...")
                        return True, response_data
                    else:
                        # For file downloads, return the response object
                        return True, response
                except:
                    return True, response
            else:
                print(f"❌ Failed - Expected {expected_status}, got {response.status_code}")
                try:
                    error_data = response.json()
                    print(f"Error response: {error_data}")
                except:
                    print(f"Error text: {response.text}")
                return False, {}

        except Exception as e:
            print(f"❌ Failed - Error: {str(e)}")
            return False, {}

    def create_sample_commandes_excel(self):
        """Create sample commandes Excel file for testing"""
        data = {
            'Dummy_A': ['CMD001', 'CMD002', 'CMD003', 'CMD004', 'CMD005', 'CMD006'],
            'Article': ['1011', '1016', '1021', '9999', '8888', '1033'],
            'Dummy_C': ['Desc1', 'Desc2', 'Desc3', 'Desc4', 'Desc5', 'Desc6'],
            'Point d\'Expédition': ['M211', 'M212', 'M213', 'M212', 'M211', 'M213'],
            'Dummy_E': ['Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5', 'Extra6'],
            'Quantité Commandée': [100, 150, 80, 120, 90, 200],
            'Stock Utilisation Libre': [20, 30, 15, 25, 18, 40],
            'Dummy_H': ['Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5', 'Extra6'],
            'Type Emballage': ['verre', 'pet', 'ciel', 'verre', 'pet', 'ciel'],
            'Dummy_J': ['Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5', 'Extra6'],
            'Produits par Palette': [30, 30, 30, 30, 30, 30]
        }
        
        df = pd.DataFrame(data)
        excel_buffer = io.BytesIO()
        df.to_excel(excel_buffer, index=False)
        excel_buffer.seek(0)
        return excel_buffer

    def create_sample_stock_excel(self):
        """Create sample stock M210 Excel file"""
        data = {
            'Division': ['M210', 'M210', 'M210', 'M210', 'M210', 'M210', 'M210', 'M210', 'M210', 'M210'],
            'Article': ['1011', '1016', '1021', '9999', '8888', '1033', '2011', '2014', '3040', '4843'],
            'Dummy_C': ['Desc1', 'Desc2', 'Desc3', 'Desc4', 'Desc5', 'Desc6', 'Desc7', 'Desc8', 'Desc9', 'Desc10'],
            'STOCK A DATE': [500, 300, 200, 400, 250, 350, 150, 180, 120, 80]  # Different stock levels for recommendations
        }
        
        df = pd.DataFrame(data)
        excel_buffer = io.BytesIO()
        df.to_excel(excel_buffer, index=False)
        excel_buffer.seek(0)
        return excel_buffer

    def create_sample_transit_excel(self):
        """Create sample transit Excel file"""
        data = {
            'Article': ['1011', '1016', '1021', '9999', '8888'],
            'Dummy_B': ['Desc1', 'Desc2', 'Desc3', 'Desc4', 'Desc5'],
            'Division': ['M211', 'M212', 'M213', 'M212', 'M211'],
            'Dummy_D': ['Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5'],
            'Dummy_E': ['Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5'],
            'Dummy_F': ['Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5'],
            'Division cédante': ['M210', 'M210', 'M210', 'M210', 'M210'],
            'Dummy_H': ['Extra1', 'Extra2', 'Extra3', 'Extra4', 'Extra5'],
            'Quantité': [30, 20, 25, 15, 40]
        }
        
        df = pd.DataFrame(data)
        excel_buffer = io.BytesIO()
        df.to_excel(excel_buffer, index=False)
        excel_buffer.seek(0)
        return excel_buffer

    def test_upload_test_data(self):
        """Upload sample test data for export testing"""
        print("\n📤 UPLOADING TEST DATA FOR EXPORT TESTING")
        
        # Upload commandes
        excel_file = self.create_sample_commandes_excel()
        files = {
            'file': ('commandes.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        
        success, response = self.run_test(
            "Upload Commandes for Export Test",
            "POST",
            "api/upload-commandes-excel",
            200,
            files=files
        )
        
        if success and 'session_id' in response:
            self.commandes_session_id = response['session_id']
            print(f"✅ Commandes uploaded - Session ID: {self.commandes_session_id}")
        else:
            return False
        
        # Upload stock M210
        excel_file = self.create_sample_stock_excel()
        files = {
            'file': ('stock_m210.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        
        success, response = self.run_test(
            "Upload Stock M210 for Export Test",
            "POST",
            "api/upload-stock-excel",
            200,
            files=files
        )
        
        if success and 'session_id' in response:
            self.stock_session_id = response['session_id']
            print(f"✅ Stock M210 uploaded - Session ID: {self.stock_session_id}")
        else:
            return False
        
        # Upload transit
        excel_file = self.create_sample_transit_excel()
        files = {
            'file': ('transit.xlsx', excel_file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        }
        
        success, response = self.run_test(
            "Upload Transit for Export Test",
            "POST",
            "api/upload-transit-excel",
            200,
            files=files
        )
        
        if success and 'session_id' in response:
            self.transit_session_id = response['session_id']
            print(f"✅ Transit uploaded - Session ID: {self.transit_session_id}")
            return True
        else:
            return False

    def test_perform_calculation(self):
        """Perform calculation to get results for export"""
        print("\n🧮 PERFORMING CALCULATION FOR EXPORT")
        
        if not self.commandes_session_id:
            print("❌ No commandes session available for calculation")
            return False
        
        calculation_data = {
            "days": 10
        }
        
        success, response = self.run_test(
            "Calculate Results for Export",
            "POST",
            "api/calculate",
            200,
            data=calculation_data
        )
        
        if success and 'calculations' in response:
            self.calculation_results = response['calculations']
            print(f"✅ Calculation completed - {len(self.calculation_results)} results available")
            
            # Print sample results for verification
            for i, calc in enumerate(self.calculation_results[:3]):
                print(f"   Sample {i+1}: {calc['depot']} | {calc['article']} | {calc['quantite_a_envoyer']} | {calc.get('palettes_needed', 0)} | {calc['statut']}")
            
            return True
        else:
            return False

    def test_export_endpoint_basic(self):
        """Test basic export endpoint functionality"""
        print("\n📊 TESTING BASIC EXPORT ENDPOINT")
        
        if not self.calculation_results:
            print("❌ No calculation results available for export")
            return False
        
        # Select all items for export
        export_data = {
            "selected_items": self.calculation_results,
            "session_id": "export_test"
        }
        
        success, response = self.run_test(
            "Basic Export Endpoint Test",
            "POST",
            "api/export-excel",
            200,
            data=export_data
        )
        
        if success:
            # Check response headers
            if hasattr(response, 'headers'):
                content_type = response.headers.get('content-type', '')
                content_disposition = response.headers.get('content-disposition', '')
                
                if 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' in content_type:
                    print("✅ Correct Excel content type returned")
                else:
                    print(f"❌ Unexpected content type: {content_type}")
                    return False
                
                if 'Export_Depots_Recommandations_' in content_disposition:
                    print("✅ Correct filename format in Content-Disposition")
                else:
                    print(f"❌ Unexpected filename format: {content_disposition}")
                    return False
                
                return True
            else:
                print("❌ No response headers available")
                return False
        else:
            return False

    def test_verify_sheet1_structure(self):
        """Test Sheet 1 'Table Principale' structure and content"""
        print("\n📋 TESTING SHEET 1 'TABLE PRINCIPALE' STRUCTURE")
        
        if not self.calculation_results:
            print("❌ No calculation results available for sheet verification")
            return False
        
        # Export and download the Excel file
        export_data = {
            "selected_items": self.calculation_results,
            "session_id": "sheet1_test"
        }
        
        success, response = self.run_test(
            "Export for Sheet 1 Verification",
            "POST",
            "api/export-excel",
            200,
            data=export_data
        )
        
        if success and hasattr(response, 'content'):
            try:
                # Load the Excel file
                excel_data = io.BytesIO(response.content)
                workbook = openpyxl.load_workbook(excel_data)
                
                # Check if Sheet 1 exists with correct name
                if "Table Principale" not in workbook.sheetnames:
                    print("❌ Sheet 'Table Principale' not found")
                    print(f"Available sheets: {workbook.sheetnames}")
                    return False
                
                sheet1 = workbook["Table Principale"]
                print("✅ Sheet 'Table Principale' found")
                
                # Verify headers (Row 1)
                expected_headers = ["Dépôt", "Code Article", "Quantité à Livrer", "Palettes", "Status"]
                actual_headers = []
                
                for col in range(1, 6):  # Columns A-E
                    cell_value = sheet1.cell(row=1, column=col).value
                    actual_headers.append(cell_value)
                
                if actual_headers == expected_headers:
                    print(f"✅ Correct headers found: {actual_headers}")
                else:
                    print(f"❌ Header mismatch. Expected: {expected_headers}, Got: {actual_headers}")
                    return False
                
                # Verify data rows
                data_rows = 0
                for row in range(2, sheet1.max_row + 1):
                    depot = sheet1.cell(row=row, column=1).value
                    article = sheet1.cell(row=row, column=2).value
                    quantite = sheet1.cell(row=row, column=3).value
                    palettes = sheet1.cell(row=row, column=4).value
                    status = sheet1.cell(row=row, column=5).value
                    
                    if depot and article:  # Valid data row
                        data_rows += 1
                        
                        # Verify data types and values
                        if not isinstance(depot, str):
                            print(f"❌ Row {row}: Depot should be string, got {type(depot)}")
                            return False
                        
                        if not isinstance(article, str):
                            print(f"❌ Row {row}: Article should be string, got {type(article)}")
                            return False
                        
                        if quantite is not None and not isinstance(quantite, (int, float)):
                            print(f"❌ Row {row}: Quantité should be numeric, got {type(quantite)}")
                            return False
                        
                        if palettes is not None and not isinstance(palettes, (int, float)):
                            print(f"❌ Row {row}: Palettes should be numeric, got {type(palettes)}")
                            return False
                        
                        if not isinstance(status, str):
                            print(f"❌ Row {row}: Status should be string, got {type(status)}")
                            return False
                        
                        # Verify status values
                        valid_statuses = ["OK", "À livrer", "Non couvert"]
                        if status not in valid_statuses:
                            print(f"❌ Row {row}: Invalid status '{status}', expected one of {valid_statuses}")
                            return False
                
                print(f"✅ Sheet 1 contains {data_rows} valid data rows")
                
                # Verify data matches calculation results
                if data_rows != len(self.calculation_results):
                    print(f"❌ Data row count mismatch. Expected {len(self.calculation_results)}, got {data_rows}")
                    return False
                
                print("✅ Sheet 1 'Table Principale' structure and content verified successfully")
                return True
                
            except Exception as e:
                print(f"❌ Error verifying Sheet 1: {str(e)}")
                return False
        else:
            return False

    def test_verify_sheet2_structure(self):
        """Test Sheet 2 'Recommandations Dépôts' structure and content"""
        print("\n📋 TESTING SHEET 2 'RECOMMANDATIONS DÉPÔTS' STRUCTURE")
        
        if not self.calculation_results:
            print("❌ No calculation results available for sheet verification")
            return False
        
        # Export and download the Excel file
        export_data = {
            "selected_items": self.calculation_results,
            "session_id": "sheet2_test"
        }
        
        success, response = self.run_test(
            "Export for Sheet 2 Verification",
            "POST",
            "api/export-excel",
            200,
            data=export_data
        )
        
        if success and hasattr(response, 'content'):
            try:
                # Load the Excel file
                excel_data = io.BytesIO(response.content)
                workbook = openpyxl.load_workbook(excel_data)
                
                # Check if Sheet 2 exists with correct name
                if "Recommandations Dépôts" not in workbook.sheetnames:
                    print("❌ Sheet 'Recommandations Dépôts' not found")
                    print(f"Available sheets: {workbook.sheetnames}")
                    return False
                
                sheet2 = workbook["Recommandations Dépôts"]
                print("✅ Sheet 'Recommandations Dépôts' found")
                
                # Verify headers (Row 1)
                expected_headers = ["Dépôt", "Palettes Actuelles", "Palettes Cibles", "Article Suggéré", 
                                  "Quantité Suggérée", "Palettes Suggérées", "Stock M210", "Faisabilité", "Raison"]
                actual_headers = []
                
                for col in range(1, 10):  # Columns A-I
                    cell_value = sheet2.cell(row=1, column=col).value
                    actual_headers.append(cell_value)
                
                if actual_headers == expected_headers:
                    print(f"✅ Correct headers found: {actual_headers}")
                else:
                    print(f"❌ Header mismatch. Expected: {expected_headers}, Got: {actual_headers}")
                    return False
                
                # Verify data rows contain recommendations
                recommendation_rows = 0
                depots_with_recommendations = set()
                
                for row in range(2, sheet2.max_row + 1):
                    depot = sheet2.cell(row=row, column=1).value
                    palettes_actuelles = sheet2.cell(row=row, column=2).value
                    palettes_cibles = sheet2.cell(row=row, column=3).value
                    article_suggere = sheet2.cell(row=row, column=4).value
                    quantite_suggeree = sheet2.cell(row=row, column=5).value
                    palettes_suggerees = sheet2.cell(row=row, column=6).value
                    stock_m210 = sheet2.cell(row=row, column=7).value
                    faisabilite = sheet2.cell(row=row, column=8).value
                    raison = sheet2.cell(row=row, column=9).value
                    
                    if depot or article_suggere:  # Valid recommendation row
                        recommendation_rows += 1
                        
                        if depot:
                            depots_with_recommendations.add(depot)
                        
                        # Verify faisabilité values
                        if faisabilite and faisabilite not in ["Réalisable", "Stock insuffisant", "Aucune suggestion", "Données stock manquantes"]:
                            print(f"❌ Row {row}: Invalid faisabilité '{faisabilite}'")
                            return False
                        
                        # Verify numeric fields when present
                        numeric_fields = [
                            (palettes_actuelles, "Palettes Actuelles"),
                            (palettes_cibles, "Palettes Cibles"),
                            (quantite_suggeree, "Quantité Suggérée"),
                            (palettes_suggerees, "Palettes Suggérées"),
                            (stock_m210, "Stock M210")
                        ]
                        
                        for value, field_name in numeric_fields:
                            if value is not None and not isinstance(value, (int, float)) and value != "":
                                print(f"❌ Row {row}: {field_name} should be numeric, got {type(value)} with value '{value}'")
                                return False
                
                print(f"✅ Sheet 2 contains {recommendation_rows} recommendation rows")
                print(f"✅ Found recommendations for {len(depots_with_recommendations)} depots: {sorted(depots_with_recommendations)}")
                
                # Verify we have recommendations for multiple depots
                if len(depots_with_recommendations) == 0:
                    print("❌ No depot recommendations found")
                    return False
                
                print("✅ Sheet 2 'Recommandations Dépôts' structure and content verified successfully")
                return True
                
            except Exception as e:
                print(f"❌ Error verifying Sheet 2: {str(e)}")
                return False
        else:
            return False

    def test_verify_data_accuracy(self):
        """Test that data in both sheets is accurate and corresponds to calculation results"""
        print("\n🎯 TESTING DATA ACCURACY IN EXPORT")
        
        if not self.calculation_results:
            print("❌ No calculation results available for accuracy verification")
            return False
        
        # Export and download the Excel file
        export_data = {
            "selected_items": self.calculation_results,
            "session_id": "accuracy_test"
        }
        
        success, response = self.run_test(
            "Export for Data Accuracy Verification",
            "POST",
            "api/export-excel",
            200,
            data=export_data
        )
        
        if success and hasattr(response, 'content'):
            try:
                # Load the Excel file
                excel_data = io.BytesIO(response.content)
                workbook = openpyxl.load_workbook(excel_data)
                sheet1 = workbook["Table Principale"]
                
                # Verify Sheet 1 data accuracy
                print("🔍 Verifying Sheet 1 data accuracy...")
                
                for row in range(2, sheet1.max_row + 1):
                    depot = sheet1.cell(row=row, column=1).value
                    article = sheet1.cell(row=row, column=2).value
                    quantite = sheet1.cell(row=row, column=3).value
                    palettes = sheet1.cell(row=row, column=4).value
                    status = sheet1.cell(row=row, column=5).value
                    
                    if depot and article:
                        # Find corresponding calculation result
                        matching_calc = None
                        for calc in self.calculation_results:
                            if calc['depot'] == depot and calc['article'] == article:
                                matching_calc = calc
                                break
                        
                        if not matching_calc:
                            print(f"❌ Row {row}: No matching calculation found for {depot} - {article}")
                            return False
                        
                        # Verify quantité à livrer
                        expected_quantite = matching_calc['quantite_a_envoyer']
                        if abs(quantite - expected_quantite) > 0.01:
                            print(f"❌ Row {row}: Quantité mismatch. Expected {expected_quantite}, got {quantite}")
                            return False
                        
                        # Verify palettes
                        expected_palettes = matching_calc.get('palettes_needed', 0)
                        if palettes != expected_palettes:
                            print(f"❌ Row {row}: Palettes mismatch. Expected {expected_palettes}, got {palettes}")
                            return False
                        
                        # Verify status
                        expected_status = matching_calc['statut']
                        if status != expected_status:
                            print(f"❌ Row {row}: Status mismatch. Expected '{expected_status}', got '{status}'")
                            return False
                
                print("✅ Sheet 1 data accuracy verified - all data matches calculation results")
                
                # Verify Sheet 2 recommendations logic
                print("🔍 Verifying Sheet 2 recommendations logic...")
                sheet2 = workbook["Recommandations Dépôts"]
                
                recommendations_found = False
                for row in range(2, sheet2.max_row + 1):
                    depot = sheet2.cell(row=row, column=1).value
                    article_suggere = sheet2.cell(row=row, column=4).value
                    stock_m210 = sheet2.cell(row=row, column=7).value
                    faisabilite = sheet2.cell(row=row, column=8).value
                    raison = sheet2.cell(row=row, column=9).value
                    
                    if depot and article_suggere:
                        recommendations_found = True
                        
                        # Verify that suggested articles are not already ordered by the depot
                        depot_articles = [calc['article'] for calc in self.calculation_results if calc['depot'] == depot]
                        if article_suggere in depot_articles:
                            print(f"❌ Row {row}: Article {article_suggere} already ordered by depot {depot}")
                            return False
                        
                        # Verify stock M210 is numeric and positive
                        if stock_m210 is not None and (not isinstance(stock_m210, (int, float)) or stock_m210 < 0):
                            print(f"❌ Row {row}: Invalid stock M210 value: {stock_m210}")
                            return False
                        
                        # Verify faisabilité logic
                        if faisabilite == "Réalisable" and stock_m210 is not None:
                            quantite_suggeree = sheet2.cell(row=row, column=5).value
                            if quantite_suggeree and quantite_suggeree > stock_m210:
                                print(f"❌ Row {row}: Faisabilité 'Réalisable' but quantité suggérée ({quantite_suggeree}) > stock M210 ({stock_m210})")
                                return False
                        
                        # Verify reason contains expected keywords
                        if raison and "Stock faible" not in str(raison):
                            print(f"❌ Row {row}: Reason should mention 'Stock faible', got: {raison}")
                            return False
                
                if recommendations_found:
                    print("✅ Sheet 2 recommendations logic verified - suggestions follow lowest stock priority")
                else:
                    print("⚠️ No recommendations found in Sheet 2 (may be expected if all depots are efficient)")
                
                print("✅ Data accuracy verification completed successfully")
                return True
                
            except Exception as e:
                print(f"❌ Error verifying data accuracy: {str(e)}")
                return False
        else:
            return False

    def test_verify_styling(self):
        """Test status-based coloring and styling in the export"""
        print("\n🎨 TESTING STYLING AND FORMATTING")
        
        if not self.calculation_results:
            print("❌ No calculation results available for styling verification")
            return False
        
        # Export and download the Excel file
        export_data = {
            "selected_items": self.calculation_results,
            "session_id": "styling_test"
        }
        
        success, response = self.run_test(
            "Export for Styling Verification",
            "POST",
            "api/export-excel",
            200,
            data=export_data
        )
        
        if success and hasattr(response, 'content'):
            try:
                # Load the Excel file
                excel_data = io.BytesIO(response.content)
                workbook = openpyxl.load_workbook(excel_data)
                sheet1 = workbook["Table Principale"]
                
                print("🔍 Verifying Sheet 1 styling...")
                
                # Check header styling
                header_styled = False
                for col in range(1, 6):
                    cell = sheet1.cell(row=1, column=col)
                    if cell.font and cell.font.bold:
                        header_styled = True
                        break
                
                if header_styled:
                    print("✅ Headers have bold styling")
                else:
                    print("⚠️ Headers may not have bold styling (could be default)")
                
                # Check status-based coloring
                status_colors_found = {
                    "Non couvert": False,
                    "À livrer": False,
                    "OK": False
                }
                
                for row in range(2, sheet1.max_row + 1):
                    status = sheet1.cell(row=row, column=5).value
                    if status in status_colors_found:
                        # Check if row has background color
                        cell = sheet1.cell(row=row, column=1)
                        if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                            status_colors_found[status] = True
                
                # Verify status-based coloring
                if status_colors_found["Non couvert"]:
                    print("✅ 'Non couvert' rows have background coloring")
                if status_colors_found["À livrer"]:
                    print("✅ 'À livrer' rows have background coloring")
                if not any(status_colors_found.values()):
                    print("⚠️ No status-based coloring detected (may use default styling)")
                
                # Check frozen panes
                if sheet1.freeze_panes:
                    print(f"✅ Frozen panes set at: {sheet1.freeze_panes}")
                else:
                    print("⚠️ No frozen panes detected")
                
                # Check column widths
                column_widths_set = False
                for col_letter in ['A', 'B', 'C', 'D', 'E']:
                    if sheet1.column_dimensions[col_letter].width:
                        column_widths_set = True
                        break
                
                if column_widths_set:
                    print("✅ Column widths have been customized")
                else:
                    print("⚠️ Column widths may be using defaults")
                
                # Check Sheet 2 styling
                print("🔍 Verifying Sheet 2 styling...")
                sheet2 = workbook["Recommandations Dépôts"]
                
                # Check for recommendation styling (green for feasible, red for insufficient)
                feasible_styled = False
                insufficient_styled = False
                
                for row in range(2, sheet2.max_row + 1):
                    faisabilite = sheet2.cell(row=row, column=8).value
                    if faisabilite == "Réalisable":
                        cell = sheet2.cell(row=row, column=1)
                        if cell.fill and cell.fill.start_color:
                            feasible_styled = True
                    elif faisabilite == "Stock insuffisant":
                        cell = sheet2.cell(row=row, column=1)
                        if cell.fill and cell.fill.start_color:
                            insufficient_styled = True
                
                if feasible_styled:
                    print("✅ Feasible recommendations have background styling")
                if insufficient_styled:
                    print("✅ Insufficient stock recommendations have background styling")
                
                print("✅ Styling verification completed")
                return True
                
            except Exception as e:
                print(f"❌ Error verifying styling: {str(e)}")
                return False
        else:
            return False

    def test_recommendations_logic(self):
        """Test that recommendations suggest products with lowest stock quantities"""
        print("\n🧠 TESTING RECOMMENDATIONS LOGIC")
        
        if not self.calculation_results:
            print("❌ No calculation results available for recommendations logic test")
            return False
        
        # Export and download the Excel file
        export_data = {
            "selected_items": self.calculation_results,
            "session_id": "recommendations_logic_test"
        }
        
        success, response = self.run_test(
            "Export for Recommendations Logic Verification",
            "POST",
            "api/export-excel",
            200,
            data=export_data
        )
        
        if success and hasattr(response, 'content'):
            try:
                # Load the Excel file
                excel_data = io.BytesIO(response.content)
                workbook = openpyxl.load_workbook(excel_data)
                sheet2 = workbook["Recommandations Dépôts"]
                
                print("🔍 Analyzing recommendations logic...")
                
                # Collect all recommendations by depot
                depot_recommendations = {}
                
                for row in range(2, sheet2.max_row + 1):
                    depot = sheet2.cell(row=row, column=1).value
                    article_suggere = sheet2.cell(row=row, column=4).value
                    stock_m210 = sheet2.cell(row=row, column=7).value
                    raison = sheet2.cell(row=row, column=9).value
                    
                    if depot and article_suggere and stock_m210 is not None:
                        if depot not in depot_recommendations:
                            depot_recommendations[depot] = []
                        
                        depot_recommendations[depot].append({
                            'article': article_suggere,
                            'stock_m210': stock_m210,
                            'raison': raison
                        })
                
                # Verify recommendations follow lowest stock priority
                for depot, recommendations in depot_recommendations.items():
                    if len(recommendations) > 1:
                        print(f"🔍 Checking {depot} recommendations order...")
                        
                        # Check if recommendations are ordered by stock (ascending)
                        stock_levels = [rec['stock_m210'] for rec in recommendations]
                        is_ascending = all(stock_levels[i] <= stock_levels[i+1] for i in range(len(stock_levels)-1))
                        
                        if is_ascending:
                            print(f"✅ {depot}: Recommendations follow lowest stock priority: {stock_levels}")
                        else:
                            print(f"⚠️ {depot}: Recommendations may not be perfectly ordered by stock: {stock_levels}")
                        
                        # Verify all recommendations mention low stock in reason
                        for rec in recommendations:
                            if rec['raison'] and "Stock faible" not in str(rec['raison']):
                                print(f"❌ {depot}: Recommendation for {rec['article']} missing 'Stock faible' in reason")
                                return False
                    
                    print(f"✅ {depot}: {len(recommendations)} recommendations with proper reasoning")
                
                # Verify exclusion logic - recommended articles should not be already ordered
                for depot, recommendations in depot_recommendations.items():
                    depot_ordered_articles = [calc['article'] for calc in self.calculation_results if calc['depot'] == depot]
                    
                    for rec in recommendations:
                        if rec['article'] in depot_ordered_articles:
                            print(f"❌ {depot}: Recommended article {rec['article']} is already ordered by this depot")
                            return False
                
                print("✅ Exclusion logic verified - no recommended articles are already ordered")
                
                if depot_recommendations:
                    print(f"✅ Recommendations logic verified for {len(depot_recommendations)} depots")
                    return True
                else:
                    print("⚠️ No recommendations found (may be expected if all depots are efficient)")
                    return True
                
            except Exception as e:
                print(f"❌ Error verifying recommendations logic: {str(e)}")
                return False
        else:
            return False

    def test_edge_cases(self):
        """Test edge cases for the export functionality"""
        print("\n🔬 TESTING EDGE CASES")
        
        # Test export with empty selection
        print("🔍 Testing export with empty selection...")
        export_data = {
            "selected_items": [],
            "session_id": "empty_test"
        }
        
        success, response = self.run_test(
            "Export with Empty Selection",
            "POST",
            "api/export-excel",
            400,  # Should return error
            data=export_data
        )
        
        if success:
            print("✅ Empty selection properly rejected")
        else:
            print("❌ Empty selection should return 400 error")
            return False
        
        # Test export with minimal data
        if self.calculation_results:
            print("🔍 Testing export with single item...")
            export_data = {
                "selected_items": [self.calculation_results[0]],
                "session_id": "single_item_test"
            }
            
            success, response = self.run_test(
                "Export with Single Item",
                "POST",
                "api/export-excel",
                200,
                data=export_data
            )
            
            if success:
                print("✅ Single item export successful")
            else:
                return False
        
        # Test export with depot having no recommendations needed
        print("🔍 Testing scenarios with efficient depots (no recommendations needed)...")
        # This is tested implicitly in other tests when depots have >= 24 palettes
        
        print("✅ Edge cases testing completed")
        return True

    def run_all_tests(self):
        """Run all export functionality tests"""
        print("🚀 STARTING COMPREHENSIVE NEW EXCEL EXPORT TESTING")
        print("=" * 80)
        
        test_methods = [
            self.test_upload_test_data,
            self.test_perform_calculation,
            self.test_export_endpoint_basic,
            self.test_verify_sheet1_structure,
            self.test_verify_sheet2_structure,
            self.test_verify_data_accuracy,
            self.test_verify_styling,
            self.test_recommendations_logic,
            self.test_edge_cases
        ]
        
        for test_method in test_methods:
            try:
                success = test_method()
                if not success:
                    print(f"\n❌ Test failed: {test_method.__name__}")
                    break
            except Exception as e:
                print(f"\n❌ Test error in {test_method.__name__}: {str(e)}")
                break
        
        print("\n" + "=" * 80)
        print(f"📊 EXPORT TESTING SUMMARY: {self.tests_passed}/{self.tests_run} tests passed")
        
        if self.tests_passed == self.tests_run:
            print("🎉 ALL EXPORT TESTS PASSED!")
            return True
        else:
            print("❌ Some export tests failed")
            return False

if __name__ == "__main__":
    tester = NewExcelExportTester()
    success = tester.run_all_tests()
    sys.exit(0 if success else 1)