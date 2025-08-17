from fastapi import FastAPI, UploadFile, File, HTTPException, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import uuid
import json
import math
from typing import List, Dict, Optional, Any

# JSON serialization helper for datetime objects
def json_serializable(obj):
    """Convert datetime objects to ISO format strings for JSON serialization"""
    if isinstance(obj, datetime):
        return obj.isoformat()
    elif isinstance(obj, dict):
        return {key: json_serializable(value) for key, value in obj.items()}
    elif isinstance(obj, list):
        return [json_serializable(item) for item in obj]
    else:
        return obj
# Lazy import for Google Generative AI is handled inside /api/chat endpoint to avoid startup failures
from pymongo import MongoClient
from pydantic import BaseModel
from fastapi.responses import StreamingResponse
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import openpyxl.comments
import openpyxl.styles

# Gemini API configuration - will be loaded lazily in chat endpoint
GEMINI_API_KEY = os.environ.get('GEMINI_API_KEY', "AIzaSyB6DG8yocaGBjwfp7TatlDDHlKELYm56BU")

app = FastAPI()

# CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# MongoDB connection
MONGO_URL = os.environ.get('MONGO_URL', 'mongodb://localhost:27017/')
try:
    client = MongoClient(MONGO_URL)
    db = client.stock_management
    collection = db.stock_data
    print("MongoDB connecté avec succès")
except Exception as e:
    print(f"Échec de la connexion MongoDB: {e}")

# Pydantic models
class CalculationRequest(BaseModel):
    days: int
    product_filter: Optional[List[str]] = None
    packaging_filter: Optional[List[str]] = None
    production_plan: Optional[List[Dict[str, Any]]] = None

class GeminiQueryRequest(BaseModel):
    query: str
    session_id: Optional[str] = None

class ChatRequest(BaseModel):
    message: str
    conversation_id: Optional[str] = None

class ExportRequest(BaseModel):
    selected_items: List[Dict]
    session_id: str

# Locally manufactured articles list
LOCALLY_MADE_ARTICLES = {
    '1011', '1016', '1021', '1022', '1033', '1040', '1051', '1059', '1069', '1071',
    '1515', '1533', '1540', '1559', '1569', '2011', '2014', '2022', '2033', '2040',
    '2069', '3040', '3043', '3056', '3140', '3149', '3156', '3249', '3256', '3948',
    '3953', '4843', '4942', '5030', '5059', '5130', '5159', '5516', '6010', '6011',
    '6016', '6020', '6030', '6040', '6059', '6069', '6120', '6140', '7435', '7436',
    '7521', '7532', '7620', '7630', '7640', '7659', '7949', '7953'
}

# Configuration storage for depot-article mapping
depot_configuration = {}

class DeliveryConfiguration(BaseModel):
    depot_article_mapping: Dict[str, List[str]]  # depot -> list of articles
    enabled: bool = True

class ConfigurationRequest(BaseModel):
    depot_article_mapping: Dict[str, List[str]]
    enabled: bool = True

# Predefined list of allowed depots for configuration
ALLOWED_DEPOTS = [
    'M112', 'M113', 'M115', 'M120', 'M130', 'M140', 'M150', 'M161', 'M170', 'M171', 
    'M212', 'M215', 'M220', 'M230', 'M240', 'M250', 'M260', 'M270', 'M280', 
    'M320', 'M330', 'M340', 'M350', 'M360', 'M362', 'M363', 'M364', 
    'M620', 'M622', 'M624', 'M626', 'M628', 'M630', 
    'M660', 'M662', 'M664', 'M666', 'M668', 'M670', 'M672', 'M674', 'M676'
]

# Store uploaded data temporarily
commandes_data = {}  # Fichier commandes
stock_data = {}      # Fichier stock M210
transit_data = {}    # Fichier transit

@app.get("/")
async def root():
    return {"message": "API de Gestion des Stocks Simplifié fonctionne"}

@app.post("/api/upload-commandes-excel")
async def upload_commandes_excel(file: UploadFile = File(...)):
    """Upload du fichier commandes avec colonnes spécifiques"""
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Veuillez télécharger un fichier Excel (.xlsx ou .xls)")
        
        # Read Excel file
        contents = await file.read()
        df = pd.read_excel(contents)
        
        # Vérifier les colonnes requises selon les spécifications - AJOUT COLONNE K
        required_columns = ['B', 'D', 'F', 'G', 'I', 'K']  # Article, Point d'Expédition, Quantité Commandée, Stock Utilisation Libre, Type Emballage, Produits par Palette
        
        # Si les colonnes sont nommées différemment, essayer de les identifier
        if 'B' not in df.columns:
            # Essayer avec les noms potentiels des colonnes
            column_mapping = {}
            if len(df.columns) >= 11:  # Au moins 11 colonnes pour avoir A-K
                column_mapping['Article'] = df.columns[1]  # Colonne B (index 1)
                column_mapping['Point d\'Expédition'] = df.columns[3]  # Colonne D (index 3) 
                column_mapping['Quantité Commandée'] = df.columns[5]  # Colonne F (index 5)
                column_mapping['Stock Utilisation Libre'] = df.columns[6]  # Colonne G (index 6)
                column_mapping['Type Emballage'] = df.columns[8]  # Colonne I (index 8)
                column_mapping['Produits par Palette'] = df.columns[10]  # Colonne K (index 10)
                
                # Renommer les colonnes pour standardiser
                df = df.rename(columns={
                    df.columns[1]: 'Article',
                    df.columns[3]: 'Point d\'Expédition', 
                    df.columns[5]: 'Quantité Commandée',
                    df.columns[6]: 'Stock Utilisation Libre',
                    df.columns[8]: 'Type Emballage',
                    df.columns[10]: 'Produits par Palette'
                })
        
        # Vérifier que nous avons les colonnes nécessaires après mapping
        required_cols = ['Article', 'Point d\'Expédition', 'Quantité Commandée', 'Stock Utilisation Libre', 'Type Emballage', 'Produits par Palette']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400, 
                detail=f"Colonnes manquantes: {', '.join(missing_cols)}. Vérifiez que le fichier contient les colonnes B, D, F, G, I, K selon les spécifications."
            )
        
        # Nettoyer et valider les données
        df['Quantité Commandée'] = pd.to_numeric(df['Quantité Commandée'], errors='coerce')
        df['Stock Utilisation Libre'] = pd.to_numeric(df['Stock Utilisation Libre'], errors='coerce')
        df['Produits par Palette'] = pd.to_numeric(df['Produits par Palette'], errors='coerce')
        
        # Nettoyer et standardiser le type d'emballage
        df['Type Emballage'] = df['Type Emballage'].astype(str).str.strip().str.lower()
        # Normaliser les valeurs d'emballage
        packaging_mapping = {
            'verre': 'verre',
            'pet': 'pet', 
            'ciel': 'ciel',
            'bottle': 'verre',  # Mapping anglais
            'plastic': 'pet',   # Mapping anglais
            'can': 'ciel'       # Mapping anglais
        }
        df['Type Emballage'] = df['Type Emballage'].map(packaging_mapping).fillna(df['Type Emballage'])
        
        # Supprimer les lignes avec données manquantes essentielles
        df = df.dropna(subset=['Article', 'Point d\'Expédition', 'Quantité Commandée', 'Type Emballage', 'Produits par Palette'])
        
        # Remplir Stock Utilisation Libre manquant par 0
        df['Stock Utilisation Libre'] = df['Stock Utilisation Libre'].fillna(0)
        
        # Validation: Produits par Palette doit être > 0
        df = df[df['Produits par Palette'] > 0]
        
        # Filtrer pour exclure M210 des dépôts destinataires (M210 ne doit jamais être approvisionné)
        df = df[df['Point d\'Expédition'] != 'M210']
        
        # Generate session ID
        session_id = str(uuid.uuid4())
        
        # Get unique values for filters
        unique_articles = sorted(df['Article'].astype(str).unique().tolist())
        unique_depots = sorted(df['Point d\'Expédition'].unique().tolist())
        unique_packaging = sorted(df['Type Emballage'].unique().tolist())
        
        # Store data
        commandes_data[session_id] = {
            'data': df.to_dict('records'),
            'upload_time': datetime.now(),
            'filters': {
                'articles': unique_articles,
                'depots': unique_depots,
                'packaging': unique_packaging
            }
        }
        
        return {
            "session_id": session_id,
            "message": "Fichier commandes uploadé avec succès",
            "summary": {
                "total_records": len(df),
                "unique_articles": len(unique_articles),
                "unique_depots": len(unique_depots),
                "unique_packaging": len(unique_packaging),
                "total_quantity": float(df['Quantité Commandée'].sum()),
                "total_stock": float(df['Stock Utilisation Libre'].sum())
            },
            "filters": {
                "articles": unique_articles,
                "depots": unique_depots,
                "packaging": unique_packaging
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lors du traitement du fichier: {str(e)}")

@app.post("/api/upload-stock-excel") 
async def upload_stock_excel(file: UploadFile = File(...)):
    """Upload du fichier stock M210 avec colonnes spécifiques"""
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Veuillez télécharger un fichier Excel (.xlsx ou .xls)")
        
        # Read Excel file
        contents = await file.read()
        df = pd.read_excel(contents)
        
        # Vérifier les colonnes selon les spécifications
        # Colonne A: Division, B: Article, D: STOCK A DATE
        if len(df.columns) >= 4:
            df = df.rename(columns={
                df.columns[0]: 'Division',
                df.columns[1]: 'Article', 
                df.columns[3]: 'STOCK A DATE'
            })
        
        required_cols = ['Division', 'Article', 'STOCK A DATE']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400, 
                detail=f"Colonnes manquantes: {', '.join(missing_cols)}. Vérifiez que le fichier contient les colonnes A, B, D selon les spécifications."
            )
        
        # Filtrer uniquement M210 (dépôt central)
        df = df[df['Division'] == 'M210']
        
        if df.empty:
            raise HTTPException(
                status_code=400, 
                detail="Aucune donnée trouvée pour le dépôt M210. Vérifiez que le fichier contient des données avec Division = M210."
            )
        
        # Nettoyer les données
        df['STOCK A DATE'] = pd.to_numeric(df['STOCK A DATE'], errors='coerce')
        df = df.dropna(subset=['Article', 'STOCK A DATE'])
        
        # Generate session ID
        session_id = str(uuid.uuid4())
        
        # Get unique articles
        unique_articles = sorted(df['Article'].astype(str).unique().tolist())
        
        # Store data
        stock_data[session_id] = {
            'data': df.to_dict('records'),
            'upload_time': datetime.now(),
            'summary': {
                'total_articles': len(unique_articles),
                'total_stock_m210': float(df['STOCK A DATE'].sum())
            }
        }
        
        return {
            "session_id": session_id,
            "message": "Fichier stock M210 uploadé avec succès",
            "summary": {
                "total_records": len(df),
                "unique_articles": len(unique_articles),
                "total_stock_m210": float(df['STOCK A DATE'].sum())
            },
            "filters": {
                "articles": sorted(unique_articles)
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lors du traitement du fichier: {str(e)}")

@app.post("/api/upload-transit-excel")
async def upload_transit_excel(file: UploadFile = File(...)):
    """Upload du fichier transit avec colonnes spécifiques"""
    try:
        # Validate file type
        if not file.filename.endswith(('.xlsx', '.xls')):
            raise HTTPException(status_code=400, detail="Veuillez télécharger un fichier Excel (.xlsx ou .xls)")
        
        # Read Excel file
        contents = await file.read()
        df = pd.read_excel(contents)
        
        # Vérifier les colonnes selon les spécifications  
        # Colonne A: Article, C: Division, G: Division cédante, I: Quantité
        if len(df.columns) >= 9:
            df = df.rename(columns={
                df.columns[0]: 'Article',
                df.columns[2]: 'Division',
                df.columns[6]: 'Division cédante',
                df.columns[8]: 'Quantité'
            })
        
        required_cols = ['Article', 'Division', 'Division cédante', 'Quantité']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            raise HTTPException(
                status_code=400, 
                detail=f"Colonnes manquantes: {', '.join(missing_cols)}. Vérifiez que le fichier contient les colonnes A, C, G, I selon les spécifications."
            )
        
        # Filtrer seulement les transits depuis M210 (dépôt central)
        df = df[df['Division cédante'] == 'M210']
        
        # Nettoyer les données
        df['Quantité'] = pd.to_numeric(df['Quantité'], errors='coerce')
        df = df.dropna(subset=['Article', 'Division', 'Quantité'])
        
        # Generate session ID
        session_id = str(uuid.uuid4())
        
        # Store data
        transit_data[session_id] = {
            'data': df.to_dict('records'),
            'upload_time': datetime.now()
        }
        
        return {
            "session_id": session_id,
            "message": "Fichier transit uploadé avec succès",
            "summary": {
                "total_records": len(df),
                "total_transit_quantity": float(df['Quantité'].sum()),
                "unique_articles": len(df['Article'].astype(str).unique()),
                "unique_destinations": len(df['Division'].unique())
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lors du traitement du fichier: {str(e)}")

@app.post("/api/calculate")
async def calculate_requirements(request: CalculationRequest):
    """Calcul simplifié avec la nouvelle formule"""
    try:
        # Vérifier qu'on a au moins les données commandes
        if not commandes_data:
            raise HTTPException(status_code=400, detail="Aucune donnée de commandes uploadée")
        
        # Prendre la dernière session de commandes uploadée
        commandes_session_id = list(commandes_data.keys())[-1]
        commandes_df = pd.DataFrame(commandes_data[commandes_session_id]['data'])
        
        # Appliquer le filtre de type d'emballage si spécifié
        if request.packaging_filter and len(request.packaging_filter) > 0:
            commandes_df = commandes_df[commandes_df['Type Emballage'].isin(request.packaging_filter)]
            
        # Appliquer la configuration depot-article si activée
        if depot_configuration.get("enabled", False) and depot_configuration.get("depot_article_mapping"):
            depot_article_mapping = depot_configuration["depot_article_mapping"]
            
            # Filtrer pour ne garder que les combinaisons depot-article configurées
            filtered_rows = []
            for _, row in commandes_df.iterrows():
                depot = str(row['Point d\'Expédition'])
                article = str(row['Article'])
                
                # Vérifier si ce dépôt est configuré et si cet article est autorisé pour ce dépôt
                if depot in depot_article_mapping and article in depot_article_mapping[depot]:
                    filtered_rows.append(row)
            
            if filtered_rows:
                commandes_df = pd.DataFrame(filtered_rows)
            else:
                # Aucune combinaison depot-article configurée trouvée
                raise HTTPException(status_code=400, detail="Aucune combinaison dépôt-article configurée trouvée dans les données")
            
        if commandes_df.empty:
            raise HTTPException(status_code=400, detail="Aucune donnée après application des filtres d'emballage et de configuration")
        
        # Obtenir les données de stock M210 si disponibles
        stock_m210 = {}
        if stock_data:
            stock_session_id = list(stock_data.keys())[-1] 
            stock_df = pd.DataFrame(stock_data[stock_session_id]['data'])
            # Créer un dictionnaire article -> stock disponible M210
            for _, row in stock_df.iterrows():
                stock_m210[str(row['Article'])] = float(row['STOCK A DATE'])
        
        # Ajouter les quantités du plan de production au stock M210
        if request.production_plan:
            for production_item in request.production_plan:
                article = str(production_item['article'])
                quantity = float(production_item['quantity'])
                current_stock = stock_m210.get(article, 0)
                stock_m210[article] = current_stock + quantity
        
        # Obtenir les données de transit si disponibles
        transit_stocks = {}
        if transit_data:
            transit_session_id = list(transit_data.keys())[-1]
            transit_df = pd.DataFrame(transit_data[transit_session_id]['data'])
            # Créer un dictionnaire (article, depot) -> quantité en transit
            for _, row in transit_df.iterrows():
                key = (str(row['Article']), str(row['Division']))
                transit_stocks[key] = transit_stocks.get(key, 0) + float(row['Quantité'])
        
        # Grouper les commandes par (Article, Point d'Expédition, Type Emballage)
        grouped = commandes_df.groupby(['Article', 'Point d\'Expédition', 'Type Emballage']).agg({
            'Quantité Commandée': 'sum',  # CQM (Consommation Quotidienne Moyenne)
            'Stock Utilisation Libre': 'first',  # Stock actuel dans le dépôt
            'Produits par Palette': 'first'  # Produits par palette pour cet article
        }).reset_index()
        
        # Calculer pour chaque ligne
        results = []
        for _, row in grouped.iterrows():
            article = str(row['Article'])
            depot = str(row['Point d\'Expédition'])
            packaging = str(row['Type Emballage'])
            cqm = float(row['Quantité Commandée'])
            stock_actuel = float(row['Stock Utilisation Libre'])
            produits_par_palette = float(row['Produits par Palette'])
            
            # Obtenir le stock en transit pour cet article vers ce dépôt
            transit_key = (article, depot)
            stock_transit = transit_stocks.get(transit_key, 0)
            
            # Appliquer la nouvelle formule
            # Quantité à Envoyer = max(0, (Quantité Commandée × Jours à Couvrir) - Stock Utilisation Libre - Quantité en Transit)
            quantite_requise = cqm * request.days
            quantite_a_envoyer = max(0, quantite_requise - stock_actuel - stock_transit)
            
            # Vérifier le stock disponible à M210
            stock_dispo_m210 = stock_m210.get(article, 0)
            
            # Check if article is locally made
            is_locally_made = str(article) in LOCALLY_MADE_ARTICLES
            sourcing_status = 'local' if is_locally_made else 'external'
            sourcing_text = 'Production Locale' if is_locally_made else 'Sourcing Externe'
            
            # Déterminer le statut
            if quantite_a_envoyer == 0:
                statut = "OK"
                statut_color = "green"
            elif quantite_a_envoyer <= stock_dispo_m210:
                statut = "À livrer"
                statut_color = "orange"
            else:
                statut = "Non couvert"
                statut_color = "red"
            
            # NOUVEAU CALCUL DES PALETTES: utiliser la valeur de colonne K pour cet article (arrondi au supérieur)
            palettes_needed = math.ceil(quantite_a_envoyer / produits_par_palette) if quantite_a_envoyer > 0 and produits_par_palette > 0 else 0
            
            # NOUVEAU: Calcul des jours de recouvrement
            # Jours de recouvrement = (Stock Actuel + Quantité en Transit) / Consommation Quotidienne Moyenne
            jours_recouvrement = 0
            if cqm > 0:  # Éviter la division par zéro
                # La Quantité Commandée représente déjà la consommation quotidienne
                cqm_daily = cqm  # Consommation quotidienne directe
                # Inclure le stock en transit dans le calcul
                stock_total_disponible = stock_actuel + stock_transit
                jours_recouvrement = round(stock_total_disponible / cqm_daily, 1)
            
            results.append({
                'article': article,
                'depot': depot,
                'packaging': packaging,
                'cqm': cqm,
                'stock_actuel': stock_actuel,
                'stock_transit': stock_transit,
                'quantite_requise': quantite_requise,
                'quantite_a_envoyer': quantite_a_envoyer,
                'stock_dispo_m210': stock_dispo_m210,
                'produits_par_palette': produits_par_palette,  # Ajouter cette info dans les résultats
                'palettes_needed': palettes_needed,
                'jours_recouvrement': jours_recouvrement,  # NOUVEAU: jours de recouvrement
                'statut': statut,
                'statut_color': statut_color,
                'sourcing_status': sourcing_status,
                'sourcing_text': sourcing_text,
                'is_locally_made': is_locally_made
            })
        
        # Calculer les statistiques de résumé
        total_items = len(results)
        items_ok = len([r for r in results if r['statut'] == 'OK'])
        items_a_livrer = len([r for r in results if r['statut'] == 'À livrer'])
        items_non_couverts = len([r for r in results if r['statut'] == 'Non couvert'])
        
        # Calculer les statistiques de sourcing
        local_items = len([r for r in results if r['is_locally_made']])
        external_items = len([r for r in results if not r['is_locally_made']])
        
        # Calculer les statistiques par dépôt (palettes et camions)
        depot_stats = {}
        for result in results:
            depot = result['depot']
            if depot not in depot_stats:
                depot_stats[depot] = {
                    'depot': depot,
                    'total_palettes': 0,
                    'total_items': 0,
                    'trucks_needed': 0,
                    'delivery_efficiency': 'Efficace'
                }
            
            depot_stats[depot]['total_palettes'] += result['palettes_needed']
            depot_stats[depot]['total_items'] += 1
        
        # Calculer le nombre de camions et l'efficacité pour chaque dépôt
        for depot in depot_stats:
            total_palettes = depot_stats[depot]['total_palettes']
            trucks_needed = math.ceil(total_palettes / 24) if total_palettes > 0 else 0
            depot_stats[depot]['trucks_needed'] = trucks_needed
            
            # Vérifier si on a des camions incomplets (pas un multiple parfait de 24)
            if total_palettes > 0 and total_palettes % 24 != 0:
                depot_stats[depot]['delivery_efficiency'] = 'Inefficace'
                # Calculer combien de palettes manquent pour compléter le dernier camion
                palettes_in_last_truck = total_palettes % 24
                depot_stats[depot]['missing_palettes'] = 24 - palettes_in_last_truck
            else:
                depot_stats[depot]['delivery_efficiency'] = 'Efficace'
                depot_stats[depot]['missing_palettes'] = 0
        
        # Convertir en liste et trier par nombre de palettes (décroissant)
        depot_summary = sorted(list(depot_stats.values()), key=lambda x: x['total_palettes'], reverse=True)
        
        # Trier les résultats par dépôt
        results_sorted = sorted(results, key=lambda x: x['depot'])
        
        return {
            "calculations": results_sorted,
            "summary": {
                "total_items": total_items,
                "items_ok": items_ok,
                "items_a_livrer": items_a_livrer,
                "items_non_couverts": items_non_couverts,
                "jours_couvrir": request.days
            },
            "sourcing_summary": {
                "local_items": local_items,
                "external_items": external_items,
                "local_percentage": round((local_items / total_items * 100) if total_items > 0 else 0, 1),
                "external_percentage": round((external_items / total_items * 100) if total_items > 0 else 0, 1)
            },
            "depot_summary": depot_summary,
            "has_stock_data": len(stock_m210) > 0,
            "has_transit_data": len(transit_stocks) > 0
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lors du calcul: {str(e)}")

@app.get("/api/sessions")
async def get_sessions():
    """Obtenir les sessions actives"""
    return {
        "commandes_sessions": list(commandes_data.keys()),
        "stock_sessions": list(stock_data.keys()),
        "transit_sessions": list(transit_data.keys())
    }

# Configuration endpoints
@app.get("/api/configuration")
async def get_configuration():
    """Get current depot-article configuration"""
    return {
        "depot_article_mapping": depot_configuration.get("depot_article_mapping", {}),
        "enabled": depot_configuration.get("enabled", False)
    }

@app.post("/api/configuration")
async def save_configuration(config: ConfigurationRequest):
    """Save depot-article configuration"""
    global depot_configuration
    depot_configuration = {
        "depot_article_mapping": config.depot_article_mapping,
        "enabled": config.enabled
    }
    return {
        "message": "Configuration sauvegardée avec succès",
        "configuration": depot_configuration
    }

@app.get("/api/available-options")
async def get_available_options():
    """Get available depots and articles from predefined depots and uploaded data"""
    # Use predefined depots list
    available_depots = set(ALLOWED_DEPOTS)
    available_articles = set()
    
    # Get articles from commandes data if available
    for session_id, session_data in commandes_data.items():
        if 'filters' in session_data:
            available_articles.update(session_data['filters'].get('articles', []))
    
    return {
        "depots": sorted(list(available_depots)),
        "articles": sorted(list(available_articles))
    }

@app.post("/api/export-excel")
async def export_excel(request: ExportRequest):
    """Export Excel avec table principale et recommandations par dépôt"""
    try:
        if not request.selected_items:
            raise HTTPException(status_code=400, detail="Aucun élément sélectionné pour l'export")
        
        # Trier les données par dépôt puis par quantité à envoyer (décroissant)
        sorted_items = sorted(request.selected_items, 
                            key=lambda x: (x['depot'], -x['quantite_a_envoyer']))
        
        # Créer un nouveau classeur Excel
        wb = openpyxl.Workbook()
        
        # ========== FEUILLE 1: TABLE PRINCIPALE ==========
        ws1 = wb.active
        ws1.title = "Table Principale"
        
        # En-têtes selon les spécifications: depot / code article / quantite a livrer / palettes / status
        headers = ["Dépôt", "Code Article", "Quantité à Livrer", "Palettes", "Status"]
        
        # Style des en-têtes
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F4F4F", end_color="4F4F4F", fill_type="solid")
        
        # Ajouter les en-têtes
        for col, header in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Ajouter les données de la table principale
        for row_idx, item in enumerate(sorted_items, 2):
            ws1.cell(row=row_idx, column=1, value=item['depot'])
            ws1.cell(row=row_idx, column=2, value=item['article'])
            ws1.cell(row=row_idx, column=3, value=item['quantite_a_envoyer'])
            ws1.cell(row=row_idx, column=4, value=item.get('palettes_needed', 0))
            ws1.cell(row=row_idx, column=5, value=item['statut'])
            
            # Style selon le statut
            if item['statut'] == 'Non couvert':
                for col in range(1, len(headers) + 1):
                    ws1.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"
                    )
            elif item['statut'] == 'À livrer':
                for col in range(1, len(headers) + 1):
                    ws1.cell(row=row_idx, column=col).fill = PatternFill(
                        start_color="FFF3E0", end_color="FFF3E0", fill_type="solid"
                    )
        
        # Ajuster les largeurs de colonne
        column_widths = [12, 15, 18, 12, 15]
        for col, width in enumerate(column_widths, 1):
            ws1.column_dimensions[get_column_letter(col)].width = width
        
        # Figer les en-têtes
        ws1.freeze_panes = 'A2'
        
        # ========== FEUILLE 2: RECOMMANDATIONS PAR DÉPÔT ==========
        ws2 = wb.create_sheet(title="Recommandations Dépôts")
        
        # En-têtes pour les recommandations
        rec_headers = ["Dépôt", "Palettes Actuelles", "Palettes Cibles", "Article Suggéré", 
                      "Quantité Suggérée", "Palettes Suggérées", "Stock M210", "Faisabilité", "Raison"]
        
        # Style des en-têtes
        for col, header in enumerate(rec_headers, 1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Générer les recommandations pour chaque dépôt unique
        depot_list = list(set(item['depot'] for item in sorted_items))
        depot_list.sort()
        
        current_row = 2
        
        # Obtenir les données nécessaires pour les suggestions
        if not commandes_data:
            # Si pas de données, ajouter une ligne d'information
            ws2.cell(row=current_row, column=1, value="Aucune donnée de commandes disponible")
            ws2.cell(row=current_row, column=2, value="pour générer des recommandations")
        else:
            # Prendre la dernière session de commandes
            commandes_session_id = list(commandes_data.keys())[-1]
            commandes_df = pd.DataFrame(commandes_data[commandes_session_id]['data'])
            
            # Obtenir les données de stock M210
            stock_m210 = {}
            if stock_data and len(stock_data) > 0:
                stock_session_id = list(stock_data.keys())[-1]
                stock_df = pd.DataFrame(stock_data[stock_session_id]['data'])
                stock_m210 = dict(zip(stock_df['Article'].astype(str), stock_df['STOCK A DATE']))
            
            # Obtenir les données de transit
            transit_stocks = {}
            if transit_data and len(transit_data) > 0:
                transit_session_id = list(transit_data.keys())[-1]
                transit_df = pd.DataFrame(transit_data[transit_session_id]['data'])
                
                for _, row in transit_df.iterrows():
                    article = str(row['Article'])
                    if article not in transit_stocks:
                        transit_stocks[article] = {}
                    depot = row['Division']
                    if depot not in transit_stocks[article]:
                        transit_stocks[article][depot] = 0
                    transit_stocks[article][depot] += row['Quantité']
            
            # Créer un dictionnaire de lookup pour les produits par palette depuis les commandes
            produits_par_palette_lookup = {}
            for _, row in commandes_df.iterrows():
                article = str(row['Article'])
                produits_par_palette_lookup[article] = float(row['Produits par Palette'])
            
            # Créer un lookup GLOBAL pour 'Produits par Palette' à partir du fichier commandes
            # Ceci permet d'utiliser la valeur de la colonne K pour chaque article, même s'il n'est pas commandé par ce dépôt spécifique
            produits_par_palette_lookup_all = {}
            try:
                for _, row in commandes_df.iterrows():
                    produits_par_palette_lookup_all[str(row['Article'])] = float(row['Produits par Palette'])
            except Exception:
                # En cas de données manquantes/impropres, laisser vide pour retomber sur le fallback 30 plus bas
                produits_par_palette_lookup_all = {}
            
            # Pour chaque dépôt, générer les recommandations
            for depot in depot_list:
                # Filtrer les commandes pour ce dépôt
                depot_commandes = commandes_df[commandes_df['Point d\'Expédition'] == depot]
                
                if depot_commandes.empty:
                    ws2.cell(row=current_row, column=1, value=depot)
                    ws2.cell(row=current_row, column=2, value="Aucune commande")
                    current_row += 1
                    continue
                
                # Calculer les palettes actuelles
                current_palettes = 0
                depot_products = []
                current_days = 10  # Valeur par défaut
                
                for _, row in depot_commandes.iterrows():
                    article = str(row['Article'])
                    packaging = row.get('Type Emballage', 'verre')
                    cqm = row['Quantité Commandée']
                    stock_actuel = row['Stock Utilisation Libre']
                    produits_par_palette = float(row['Produits par Palette'])  # Removed fallback to 30, use actual value from column K
                    stock_transit = transit_stocks.get(article, {}).get(depot, 0)
                    
                    # Calculer avec la formule
                    quantite_requise = cqm * current_days
                    quantite_a_envoyer = max(0, quantite_requise - stock_actuel - stock_transit)
                    palettes_needed = math.ceil(quantite_a_envoyer / produits_par_palette) if quantite_a_envoyer > 0 and produits_par_palette > 0 else 0
                    
                    current_palettes += palettes_needed
                    depot_products.append({'article': article, 'packaging': packaging, 'produits_par_palette': produits_par_palette})
                
                # Calculer les palettes cibles (multiples de 24)
                current_trucks = math.ceil(current_palettes / 24) if current_palettes > 0 else 1
                target_palettes = current_trucks * 24
                palettes_to_add = target_palettes - current_palettes
                
                # Nouvelle logique de recommandations: suggérer UNIQUEMENT les articles déjà envoyés à ce dépôt
                # et privilégier ceux avec le PLUS GRAND stock disponible à M210
                if palettes_to_add > 0 and stock_m210:
                    # Créer la liste des candidats à partir des produits déjà commandés pour ce dépôt
                    candidate_products = []
                    for product in depot_products:
                        article = product['article']
                        stock_quantity = stock_m210.get(article, 0)
                        if stock_quantity > 0:
                            candidate_products.append({
                                'article': article,
                                'stock_m210': stock_quantity,
                                'packaging': product.get('packaging', 'verre'),
                                'produits_par_palette': product.get('produits_par_palette') or produits_par_palette_lookup_all.get(article, 30)
                            })
                    
                    # Trier par stock décroissant (plus grand en premier)
                    candidate_products.sort(key=lambda x: x['stock_m210'], reverse=True)
                    
                    remaining_palettes = palettes_to_add
                    suggestions_count = 0
                    
                    for product in candidate_products:
                        if remaining_palettes <= 0 or suggestions_count >= 5:
                            break
                        
                        suggested_palettes = min(3, remaining_palettes)
                        products_needed_per_palette = float(product['produits_par_palette'])
                        suggested_quantity = int(suggested_palettes * products_needed_per_palette)
                        stock_available = float(product['stock_m210'])
                        can_fulfill = suggested_quantity <= stock_available
                        
                        # Ajouter une ligne de recommandation
                        if suggestions_count == 0:  # Première suggestion pour ce dépôt
                            ws2.cell(row=current_row, column=1, value=depot)
                            ws2.cell(row=current_row, column=2, value=current_palettes)
                            ws2.cell(row=current_row, column=3, value=target_palettes)
                        else:
                            ws2.cell(row=current_row, column=1, value="")
                            ws2.cell(row=current_row, column=2, value="")
                            ws2.cell(row=current_row, column=3, value="")
                        
                        ws2.cell(row=current_row, column=4, value=product['article'])
                        ws2.cell(row=current_row, column=5, value=suggested_quantity)
                        ws2.cell(row=current_row, column=6, value=suggested_palettes)
                        ws2.cell(row=current_row, column=7, value=stock_available)
                        ws2.cell(row=current_row, column=8, value='Réalisable' if can_fulfill else 'Stock insuffisant')
                        ws2.cell(row=current_row, column=9, value=f'Stock élevé ({int(stock_available)} unités)')
                        
                        # Style selon la faisabilité
                        if can_fulfill:
                            for col in range(1, len(rec_headers) + 1):
                                ws2.cell(row=current_row, column=col).fill = PatternFill(
                                    start_color="E8F5E8", end_color="E8F5E8", fill_type="solid"
                                )
                        else:
                            for col in range(1, len(rec_headers) + 1):
                                ws2.cell(row=current_row, column=col).fill = PatternFill(
                                    start_color="FFEBEE", end_color="FFEBEE", fill_type="solid"
                                )
                        
                        current_row += 1
                        remaining_palettes -= suggested_palettes
                        suggestions_count += 1
                
                else:
                    # Dépôt sans besoin de suggestions ou sans données stock
                    ws2.cell(row=current_row, column=1, value=depot)
                    ws2.cell(row=current_row, column=2, value=current_palettes)
                    ws2.cell(row=current_row, column=3, value=target_palettes)
                    ws2.cell(row=current_row, column=4, value="Aucune suggestion" if palettes_to_add <= 0 else "Données stock manquantes")
                    current_row += 1
                
                # Ligne vide entre les dépôts
                current_row += 1
        
        # Ajuster les largeurs de colonne pour les recommandations
        rec_column_widths = [12, 16, 14, 15, 16, 16, 12, 15, 25]
        for col, width in enumerate(rec_column_widths, 1):
            ws2.column_dimensions[get_column_letter(col)].width = width
        
        # Figer les en-têtes
        ws2.freeze_panes = 'A2'
        
        # Sauvegarder dans un buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Nom de fichier
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"Export_Depots_Recommandations_{timestamp}.xlsx"
        
        return StreamingResponse(
            BytesIO(output.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lors de l'export: {str(e)}")

@app.post("/api/depot-suggestions")
async def get_depot_suggestions(request: dict):
    """Génère des suggestions pour compléter 24 palettes par camion pour un dépôt"""
    try:
        depot_name = request.get('depot_name')
        current_days = request.get('days', 10)
        
        if not depot_name:
            raise HTTPException(status_code=400, detail="Nom de dépôt requis")
        
        # Vérifier qu'on a des données de commandes
        if not commandes_data:
            raise HTTPException(status_code=400, detail="Aucune donnée de commandes uploadée")
        
        # Prendre la dernière session de commandes uploadée
        commandes_session_id = list(commandes_data.keys())[-1]
        commandes_df = pd.DataFrame(commandes_data[commandes_session_id]['data'])
        
        # Construire un lookup GLOBAL pour 'Produits par Palette' à partir du fichier commandes
        # Ceci permet d'utiliser la valeur de la colonne K pour chaque article, même s'il n'est pas commandé par ce dépôt spécifique
        produits_par_palette_lookup_all = {}
        try:
            for _, row in commandes_df.iterrows():
                produits_par_palette_lookup_all[str(row['Article'])] = float(row['Produits par Palette'])
        except Exception:
            # En cas de données manquantes/impropres, laisser vide pour retomber sur le fallback 30 plus bas
            produits_par_palette_lookup_all = {}
        
        # Obtenir les données de stock et transit actuelles
        stock_m210 = {}
        if stock_data and len(stock_data) > 0:
            stock_session_id = list(stock_data.keys())[-1]
            stock_df = pd.DataFrame(stock_data[stock_session_id]['data'])
            stock_m210 = dict(zip(stock_df['Article'].astype(str), stock_df['STOCK A DATE']))
        
        transit_stocks = {}
        if transit_data and len(transit_data) > 0:
            transit_session_id = list(transit_data.keys())[-1]
            transit_df = pd.DataFrame(transit_data[transit_session_id]['data'])
            
            for _, row in transit_df.iterrows():
                article = str(row['Article'])
                if article not in transit_stocks:
                    transit_stocks[article] = {}
                depot = row['Division']
                if depot not in transit_stocks[article]:
                    transit_stocks[article][depot] = 0
                transit_stocks[article][depot] += row['Quantité']
        
        # Filtrer les commandes pour ce dépôt spécifique
        depot_commandes = commandes_df[commandes_df['Point d\'Expédition'] == depot_name]
        
        if depot_commandes.empty:
            return {
                "depot_name": depot_name,
                "current_palettes": 0,
                "target_palettes": 24,
                "suggestions": [],
                "message": f"Aucune commande trouvée pour le dépôt {depot_name}"
            }
        
        # Calculer les palettes actuelles pour ce dépôt
        current_palettes = 0
        depot_products = []
        
        for _, row in depot_commandes.iterrows():
            article = str(row['Article'])
            depot = row['Point d\'Expédition']
            packaging = row['Type Emballage']
            cqm = row['Quantité Commandée']
            stock_actuel = row['Stock Utilisation Libre']
            produits_par_palette = float(row['Produits par Palette'])
            stock_transit = transit_stocks.get(article, {}).get(depot, 0)
            
            # Calculer avec la formule actuelle
            quantite_requise = cqm * current_days
            quantite_a_envoyer = max(0, quantite_requise - stock_actuel - stock_transit)
            palettes_needed = math.ceil(quantite_a_envoyer / produits_par_palette) if quantite_a_envoyer > 0 and produits_par_palette > 0 else 0
            
            current_palettes += palettes_needed
            
            depot_products.append({
                'article': article,
                'packaging': packaging,
                'cqm': cqm,
                'stock_actuel': stock_actuel,
                'stock_transit': stock_transit,
                'quantite_a_envoyer': quantite_a_envoyer,
                'palettes_needed': palettes_needed,
                'produits_par_palette': produits_par_palette,
                'stock_dispo_m210': stock_m210.get(article, 0)
            })
        
        # Calculer combien de palettes sont nécessaires pour atteindre des multiples de 24
        current_trucks = math.ceil(current_palettes / 24) if current_palettes > 0 else 1
        target_palettes = current_trucks * 24
        palettes_to_add = target_palettes - current_palettes
        
        suggestions = []
        if palettes_to_add > 0:
            if not stock_m210:
                suggestions.append({
                    'message': 'Aucune donnée de stock M210 disponible pour générer des suggestions'
                })
            else:
                # Suggérer UNIQUEMENT les articles déjà envoyés à ce dépôt et classer par plus grand stock M210
                # Construire la liste des candidats depuis depot_products
                candidate_products = []
                for product in depot_products:
                    article = product['article']
                    stock_quantity = stock_m210.get(article, 0)
                    if stock_quantity > 0:
                        candidate_products.append({
                            'article': article,
                            'packaging': product.get('packaging', 'verre'),
                            'stock_m210': stock_quantity,
                            'produits_par_palette': product.get('produits_par_palette') or produits_par_palette_lookup_all.get(article, 30)
                        })
                
                # Trier par stock M210 décroissant (plus grand d'abord)
                candidate_products.sort(key=lambda x: x['stock_m210'], reverse=True)
                
                remaining_palettes = palettes_to_add
                
                for product in candidate_products:
                    if remaining_palettes <= 0:
                        break
                    
                    suggested_palettes = min(3, remaining_palettes)
                    products_needed_per_palette = float(product['produits_par_palette'])
                    suggested_quantity = int(suggested_palettes * products_needed_per_palette)
                    stock_available = float(product['stock_m210'])
                    can_fulfill = suggested_quantity <= stock_available
                    
                    suggestions.append({
                        'article': product['article'],
                        'packaging': product['packaging'],
                        'stock_m210': stock_available,
                        'suggested_quantity': suggested_quantity,
                        'suggested_palettes': suggested_palettes,
                        'can_fulfill': can_fulfill,
                        'feasibility': 'Réalisable' if can_fulfill else 'Stock insuffisant',
                        'reason': f'Stock élevé ({int(stock_available)} unités) - Priorité à l\'envoi'
                    })
                    
                    remaining_palettes -= suggested_palettes
        
        return {
            "depot_name": depot_name,
            "current_palettes": current_palettes,
            "current_trucks": current_trucks,
            "target_palettes": target_palettes,
            "palettes_to_add": palettes_to_add,
            "suggestions": suggestions[:5],  # Limiter à 5 suggestions max
            "efficiency_status": "Efficace" if current_palettes > 0 and current_palettes % 24 == 0 else "Inefficace"
        }
        
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erreur lors de la génération des suggestions: {str(e)}")

@app.post("/api/chat")
async def chat_with_ai(request: ChatRequest):
    """Chat with AI about uploaded files and inventory data"""
    try:
        # Helper: minimal fallback response (no external API)
        def minimal_bullets(context: dict) -> str:
            cmd = context.get('commandes', {}).get('total_records', 0)
            stk = context.get('stock', {}).get('total_records', 0)
            trn = context.get('transit', {}).get('total_records', 0)
            return f"* Commandes: {cmd}\n* Stock: {stk}\n* Transit: {trn}"
        
        # Get available data context
        data_context = {}
        
        # Add commandes data if available
        if commandes_data:
            latest_commandes = list(commandes_data.keys())[-1]
            commandes_info = commandes_data[latest_commandes]
            data_context['commandes'] = {
                'total_records': len(commandes_info['data']),
                'summary': commandes_info.get('summary', {}),
                'sample_data': commandes_info['data'][:3] if len(commandes_info['data']) > 0 else []
            }
        
        # Add stock data if available
        if stock_data:
            latest_stock = list(stock_data.keys())[-1]
            stock_info = stock_data[latest_stock]
            data_context['stock'] = {
                'total_records': len(stock_info['data']),
                'summary': stock_info.get('summary', {}),
                'sample_data': stock_info['data'][:3] if len(stock_info['data']) > 0 else []
            }
        
        # Add transit data if available
        if transit_data:
            latest_transit = list(transit_data.keys())[-1]
            transit_info = transit_data[latest_transit]
            data_context['transit'] = {
                'total_records': len(transit_info['data']),
                'sample_data': transit_info['data'][:3] if len(transit_info['data']) > 0 else []
            }
        
        # Build context prompt - MINIMAL BUT USEFUL
        system_prompt = """Assistant inventaire - RÉPONSES COURTES EN BULLET POINTS.

        DONNÉES DISPONIBLES:
        """
        
        if data_context:
            for data_type, info in data_context.items():
                system_prompt += f"\n{data_type.upper()}: {info['total_records']} enregistrements"
                if info['sample_data']:
                    # Provide sample data for analysis
                    safe_sample_data = json_serializable(info['sample_data'][:2])
                    system_prompt += f"\nExemples: {json.dumps(safe_sample_data, ensure_ascii=False)}"
        else:
            system_prompt += "\nAucune donnée uploadée."
        
        system_prompt += """

        INSTRUCTIONS:
        • Format bullet points OBLIGATOIRE  
        • Maximum 3 points par réponse
        • Réponses précises avec chiffres exacts
        • Analyse les données quand disponibles
        • Réponds toujours même si données limitées"""
        
        # Try Gemini; on any failure return minimal bullets instead of 500
        try:
            import google.generativeai as genai
            if not GEMINI_API_KEY:
                # No key → fallback
                response_text = minimal_bullets(data_context)
            else:
                genai.configure(api_key=GEMINI_API_KEY)
                model = genai.GenerativeModel('gemini-1.5-flash')
                full_prompt = f"{system_prompt}\n\nQUESTION UTILISATEUR: {request.message}"
                resp = model.generate_content(full_prompt)
                response_text = resp.text if getattr(resp, 'text', None) else minimal_bullets(data_context)
        except Exception:
            # ImportError or API error → fallback
            response_text = minimal_bullets(data_context)
        
        conversation_id = request.conversation_id or str(uuid.uuid4())
        return {
            "response": response_text,
            "conversation_id": conversation_id,
            "has_data": len(data_context) > 0,
            "data_types": list(data_context.keys()),
            "message": request.message
        }
        
    except Exception as e:
        # As last resort, still avoid hard 500 but give a minimal response
        return {
            "response": "* Commandes: 0\n* Stock: 0\n* Transit: 0",
            "conversation_id": request.conversation_id or str(uuid.uuid4()),
            "has_data": False,
            "data_types": [],
            "message": str(e)
        }

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8001)